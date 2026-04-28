from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Request, Body, Depends
from typing import List
import json
from pathlib import Path
import io
import logging
from sqlalchemy.orm import Session
from app.database import get_db, SessionLocal
from app.models import KeywordAttribute, User, UserKeyword

logger = logging.getLogger(__name__)

router = APIRouter()

# 配置文件路径
CONFIG_PATH = Path(__file__).parent.parent / "scraper" / "scraper_config.json"


def load_config():
    """加载配置文件"""
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"keywords": []}


def save_config(config):
    """保存配置文件"""
    with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4, ensure_ascii=False)


@router.get("/keywords")
def get_keywords():
    """获取关键词列表"""
    config = load_config()
    return config.get("keywords", [])

# json 版本接口，兼容前端 keywords.js 的调用方式
@router.post("/keywords")
def add_keyword(keyword: str = Query(..., description="要添加的关键词")):
    """添加关键词"""
    config = load_config()
    keywords = config.get("keywords", [])
    
    if keyword not in keywords:
        keywords.append(keyword)
        config["keywords"] = keywords
        save_config(config)
    
    return {"keywords": keywords}


@router.delete("/keywords")
def delete_keyword(keyword: str = Query(..., description="要删除的关键词")):
    """删除关键词"""
    config = load_config()
    keywords = config.get("keywords", [])
    
    if keyword not in keywords:
        raise HTTPException(status_code=404, detail="关键词不存在")
    
    keywords.remove(keyword)
    config["keywords"] = keywords
    save_config(config)
    
    return {"keywords": keywords}


@router.put("/keywords")
def update_keywords(keywords: List[str]):
    """批量更新关键词"""
    config = load_config()
    config["keywords"] = keywords
    save_config(config)
    
    return {"keywords": keywords}


# ========== 标签接口（路径参数格式） ==========
@router.get("/keywords/{keyword}/tags")
def get_keyword_tags(keyword: str):
    """获取关键词的标签"""
    config = load_config()
    return config.get("keyword_tags", {}).get(keyword, [])


@router.put("/keywords/{keyword}/tags")
def update_keyword_tags(keyword: str, tags: List[str] = Body(...)):
    """更新关键词的标签"""
    config = load_config()
    if "keyword_tags" not in config:
        config["keyword_tags"] = {}
    config["keyword_tags"][keyword] = tags
    save_config(config)
    return tags


# ========== 节日接口（路径参数格式） ==========
@router.get("/keywords/{keyword}/festival")
def get_keyword_festival(keyword: str):
    """获取关键词的节日"""
    config = load_config()
    return config.get("keyword_festivals", {}).get(keyword, "")


@router.put("/keywords/{keyword}/festival")
def update_keyword_festival(keyword: str, festival: str = Body(...)):
    """更新关键词的节日"""
    config = load_config()
    if "keyword_festivals" not in config:
        config["keyword_festivals"] = {}
    config["keyword_festivals"][keyword] = festival
    save_config(config)
    return festival


# ========== 大/小节日接口（路径参数格式） ==========
@router.get("/keywords/{keyword}/festival-type")
def get_keyword_festival_type(keyword: str):
    """获取关键词的大/小节日"""
    config = load_config()
    return config.get("keyword_festival_types", {}).get(keyword, "")


@router.put("/keywords/{keyword}/festival-type")
def update_keyword_festival_type(keyword: str, festival_type: str = Body(...)):
    """更新关键词的大/小节日"""
    config = load_config()
    if "keyword_festival_types" not in config:
        config["keyword_festival_types"] = {}
    config["keyword_festival_types"][keyword] = festival_type
    save_config(config)
    return festival_type


# ========== 热卖期接口（路径参数格式） ==========
@router.get("/keywords/{keyword}/hot-season")
def get_keyword_hot_season(keyword: str):
    """获取关键词的热卖期"""
    config = load_config()
    return config.get("keyword_hot_seasons", {}).get(keyword, "")


@router.put("/keywords/{keyword}/hot-season")
def update_keyword_hot_season(keyword: str, hot_season: str = Body(...)):
    """更新关键词的热卖期"""
    config = load_config()
    if "keyword_hot_seasons" not in config:
        config["keyword_hot_seasons"] = {}
    config["keyword_hot_seasons"][keyword] = hot_season
    save_config(config)
    return hot_season


# ========== 批量导入接口 ==========
@router.post("/keywords/import-with-user")
async def import_keywords_with_user(file: UploadFile = File(...)):
    try:
        import openpyxl
        from app.database import SessionLocal
        from app.models import User, UserKeyword
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        invalid = [i+2 for i, r in enumerate(rows) if len(r) < 1 or not r[0]]
        if invalid:
            raise HTTPException(status_code=400, detail=f"格式错误：第 {invalid} 行缺少关键词，第一列必须填写关键词")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {e}")

    db = SessionLocal()
    try:
        config = load_config()
        existing_kws = set(config.get("keywords", []))
        added_kws = 0
        added_relations = 0
        
        if "keyword_tags" not in config:
            config["keyword_tags"] = {}
        if "keyword_festivals" not in config:
            config["keyword_festivals"] = {}
        if "keyword_festival_types" not in config:
            config["keyword_festival_types"] = {}
        if "keyword_hot_seasons" not in config:
            config["keyword_hot_seasons"] = {}

        for row in rows:
            keyword = str(row[0]).strip() if row[0] else ""
            user_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            tags_str = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            festival = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            festival_type = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            hot_season = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            
            if not keyword:
                continue
            
            if keyword not in existing_kws:
                existing_kws.add(keyword)
                added_kws += 1
            
            if tags_str:
                tags = [t.strip() for t in tags_str.split(',') if t.strip()]
                if tags:
                    config["keyword_tags"][keyword] = tags
            
            if festival:
                config["keyword_festivals"][keyword] = festival
            
            if festival_type and festival_type in ['大节日', '小节日']:
                config["keyword_festival_types"][keyword] = festival_type
            
            if hot_season:
                config["keyword_hot_seasons"][keyword] = hot_season
            
            if user_name:
                users = [u.strip() for u in user_name.split(',') if u.strip()]
                for u_name in users:
                    user = db.query(User).filter(User.name == u_name).first()
                    if not user:
                        user = User(name=u_name)
                        db.add(user)
                        db.flush()
                    exists = db.query(UserKeyword).filter(
                        UserKeyword.user_id == user.id, 
                        UserKeyword.keyword == keyword
                    ).first()
                    if not exists:
                        db.add(UserKeyword(user_id=user.id, keyword=keyword))
                        added_relations += 1
        
        config["keywords"] = list(existing_kws)
        save_config(config)
        db.commit()
        
        return {
            "imported_keywords": added_kws,
            "imported_relations": added_relations
        }
    except Exception as e:
        db.rollback()
        logger.error(f"导入失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")
    finally:
        db.close()


@router.post("/keywords/import")
async def import_keywords(file: UploadFile = File(...)):
    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        new_kws = [str(row[0].value).strip() for row in ws.iter_rows(min_row=2) if row[0].value]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {e}")

    config = load_config()
    existing = config.get("keywords", [])
    added = [kw for kw in new_kws if kw and kw not in existing]
    config["keywords"] = existing + added
    save_config(config)
    return {"imported": len(added), "total": len(config["keywords"])}

# ========== 在 keywords.py 现有的路径参数接口后面添加以下代码 ==========

# ========== 查询参数版本（兼容前端 keywords.js） ==========

@router.get("/keywords/tags")
def get_keyword_tags_query(keyword: str = Query(...)):
    """获取关键词的标签（查询参数版本）"""
    config = load_config()
    return config.get("keyword_tags", {}).get(keyword, [])

@router.put("/keywords/tags")
def update_keyword_tags_query(keyword: str = Query(...), tags: List[str] = Body(...)):
    """更新关键词的标签（查询参数版本）"""
    config = load_config()
    if "keyword_tags" not in config:
        config["keyword_tags"] = {}
    config["keyword_tags"][keyword] = tags
    save_config(config)
    return tags

@router.get("/keywords/festival")
def get_keyword_festival_query(keyword: str = Query(...)):
    """获取关键词的节日（查询参数版本）"""
    config = load_config()
    return config.get("keyword_festivals", {}).get(keyword, "")

@router.put("/keywords/festival")
def update_keyword_festival_query(keyword: str = Query(...), festival: str = Body(...)):
    """更新关键词的节日（查询参数版本）"""
    config = load_config()
    if "keyword_festivals" not in config:
        config["keyword_festivals"] = {}
    config["keyword_festivals"][keyword] = festival
    save_config(config)
    return festival

@router.get("/keywords/festival-type")
def get_keyword_festival_type_query(keyword: str = Query(...)):
    """获取关键词的大/小节日（查询参数版本）"""
    config = load_config()
    return config.get("keyword_festival_types", {}).get(keyword, "")

@router.put("/keywords/festival-type")
def update_keyword_festival_type_query(keyword: str = Query(...), festival_type: str = Body(...)):
    """更新关键词的大/小节日（查询参数版本）"""
    config = load_config()
    if "keyword_festival_types" not in config:
        config["keyword_festival_types"] = {}
    config["keyword_festival_types"][keyword] = festival_type
    save_config(config)
    return festival_type

@router.get("/keywords/hot-season")
def get_keyword_hot_season_query(keyword: str = Query(...)):
    """获取关键词的热卖期（查询参数版本）"""
    config = load_config()
    return config.get("keyword_hot_seasons", {}).get(keyword, "")

@router.put("/keywords/hot-season")
def update_keyword_hot_season_query(keyword: str = Query(...), hot_season: str = Body(...)):
    """更新关键词的热卖期（查询参数版本）"""
    config = load_config()
    if "keyword_hot_seasons" not in config:
        config["keyword_hot_seasons"] = {}
    config["keyword_hot_seasons"][keyword] = hot_season
    save_config(config)
    return hot_season